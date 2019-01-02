import requests
import re
import sys
import csv
import argparse
import getch

from bs4 import BeautifulSoup
from lxml.html.soupparser import unescape
from openpyxl import Workbook


def parse_command_line():
    parser = argparse.ArgumentParser(description="jwxt-csv: Grade export tool")
    parser.add_argument("-u", metavar="<Username>", dest="usr", type=int, required=False,
                        help="CAS Username.")

    parser.add_argument("-p", metavar="<Password>", dest="pwd", type=str, required=False,
                        help="CAS password.")

    args = parser.parse_args()
    if args.usr is None and args.pwd is None:
        return interact_get_params()
    elif args.usr is None or args.pwd is None:
        print('Wrong arguments, please check them.')
        sys.exit(1)
    return args.usr, args.pwd


def interact_get_params():
    usr = input("CAS Username: ")
    pwd = input("CAS password: ")
    return usr, pwd


home = "http://jwxt.sustc.edu.cn/jsxsd/"
grade_site = "http://jwxt.sustc.edu.cn/jsxsd/kscj/cjcx_list"


class SUSTech:
    """
    this code is to get Sakai page for SUSTC students, they can get necessary
    information such as course slices or assignments from this modual
    """

    def __init__(self, username, password, site):

        """
        to init CAS, username and password is in need
        """
        self.site = site
        self.headers = {
            "User-Agent": 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_3)' +
                          ' AppleWebKit/537.36 (KHTML, like Gecko) Chrome/' +
                          '56.0.2924.87 Safari/537.36'}
        self.data = {
            'username': str(username),
            'password': str(password),
            '_eventId': 'submit',
        }
        self.url = 'https://cas.sustc.edu.cn/cas/login?service=' + site
        self.s = requests.session()
        r = self.s.get(self.url, headers=self.headers)
        content = r.content.decode('utf-8')
        self.data['execution'] = self._get_execution(content)
        self.loggedIn = False

    def _get_execution(self, content):
        formula = '<input.*?name="execution".*?value="(.*?)"/>'
        pattern = re.compile(formula)
        return re.findall(pattern, content)[0]

    def login(self):
        self.s.post(self.url, self.data)
        text = self._get_home_page()
        self.loggedIn = 'CAS' not in text
        return self.loggedIn

    def _check_logged(self):
        if not self.loggedIn:
            print('not logged in, permission denied')
        return self.loggedIn

    def check_logged(self):
        if not self.loggedIn:
            print('##################################################################\n\n'
                  'Fail to log in! Please check the CAS username or password.\n\n'
                  '##################################################################\n')
            getch.pause_exit(1, 'Press Any Key To Exit.')

    def _get_home_page(self):
        r = self.s.get(self.site)
        text = r.content.decode('utf-8')
        txt = unescape(text)
        return txt

    def get_home_page(self):
        if not self._check_logged():
            return
        return self._get_home_page()

    def get_home_soup(self):
        if not self.loggedIn:
            raise Exception('not logged in yet!')
        r = self.s.get(self.site)
        soup = BeautifulSoup(r.text, 'lxml')
        return soup

    def get_cookies(self):
        return self.s.cookies

    def get_website(self, url, paras=None):
        if not self.loggedIn:
            return
        r = self.s.get(url, params=paras)
        # print(r.url)
        return r.text

    def post_website(self, url, post_data):
        r = self.s.post(url, data=post_data)
        return r.text

    def get(self, *args):
        return self.get_website(*args)

    def post(self, *args):
        return self.post_website(*args)


def get_course_grade(tr):
    soup = tr.find_all('td')
    data = list()
    for i in range(len(soup)):
        data.append(soup[i].get_text())
    return data


def dump_csv(file_name, header, data):
    with open(file_name, "w", encoding="utf_8_sig", newline='') as f:
        f_csv = csv.writer(f)
        f_csv.writerow(header)
        f_csv.writerows(data)


def load_csv(file_name):
    result = []
    with open(file_name, "r", encoding="utf_8_sig") as f:
        reader = csv.reader(f)
        for item in reader:
            result.append(item)
    return result


def try2digit(a):
    for i in range(len(a)):
        for j in range(len(a[i])):
            try:
                tmp = float(a[i][j])
                a[i][j] = tmp
                continue
            except ValueError:
                pass

            try:
                tmp = int(a[i][j])
                a[i][j] = tmp
                continue
            except ValueError:
                pass
    return a


if __name__ == '__main__':
    usr, pwd = parse_command_line()
    spider = SUSTech(usr, pwd, home)
    spider.login()
    spider.check_logged()

    content = spider.get_website(grade_site)
    soup = BeautifulSoup(content, 'lxml')
    table = soup.find_all('table', attrs={'id': True})[0]
    trs = table.find_all('tr')

    # Get table header
    header_soup = trs[0].find_all('th')
    header = list()
    for i in header_soup:
        content = i.get_text()
        if "/" in content:
            header.extend(content.split("/"))
        else:
            header.append(content)
    # print(header)

    # Get table data
    data = list()
    for i in range(1, len(trs)):
        tmp = get_course_grade(trs[i])
        data.append(tmp)
        # print(tmp)
    data = try2digit(data)

    # Save as CSV
    file_name = "{}.csv".format(usr)
    dump_csv(file_name, header, data)

    # Save as xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = "Grade"
    ws.append(header)
    for d in data:
        ws.append(d)

    func1 = "=VLOOKUP(F{},'GPA lookup'!A:C,{},FALSE)"
    func2 = "=G{}*M{}"

    for i in range(2, len(data) + 2):
        cell = 'M{}'.format(i)
        ws[cell] = func1.format(i, 2)
        cell = 'N{}'.format(i)
        ws[cell] = func1.format(i, 3)

    # LOOKUP table
    ws2 = wb.create_sheet(title="GPA lookup")
    lookup = load_csv('GPAlookup.csv')
    lookup = try2digit(lookup)
    for d in lookup:
        ws2.append(d)
    wb.save("{}.xlsx".format(usr))

    print("##################################################################\n\n"
          "Successful output your grade at {}.csv and {}.xlsx.\n\n"
          "##################################################################\n".format(usr, usr))
    getch.pause_exit(0, 'Press Any Key To Exit.')
